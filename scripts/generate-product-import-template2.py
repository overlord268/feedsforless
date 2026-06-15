#!/usr/bin/env python3
"""
Generate FeedsForLess product import Excel template v2.
Run: python scripts/generate-product-import-template2.py

Master sheets are filled first. Product-only data lives in the PRODUCTS sheet.
See docs/instruction.md for encoding rules.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "product-import-template-v2-feedsforless.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1B4332")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DESC_FILL = PatternFill("solid", fgColor="D8F3DC")
DESC_FONT = Font(italic=True, color="495057", size=9)
THIN = Side(style="thin", color="ADB5BD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_desc_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = DESC_FILL
        cell.font = DESC_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER


def auto_width(ws, min_width=12, max_width=65):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = min_width
        for cell in column_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = length


def add_data_sheet(wb, name, headers, descriptions, examples=None):
    ws = wb.create_sheet(name)
    for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=2, column=col, value=desc)
    style_header_row(ws, 1, len(headers))
    style_desc_row(ws, 2, len(headers))
    if examples:
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A3"
    auto_width(ws)
    return ws


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # --- Master sheets (fill first — see instruction.md) ---

    add_data_sheet(
        wb, "CATEGORIES",
        ["label", "slug", "parent_slug"],
        [
            "REQUIRED — Display name",
            "REQUIRED — Unique slug (lowercase, hyphens)",
            "OPTIONAL — Parent category slug. Empty = root. Parent must exist first.",
        ],
        [
            ["Phosphates", "phosphates", ""],
            ["Feed Additives", "feed-additives", ""],
            ["Minerals", "minerals", "feed-additives"],
            ["Amino Acids", "amino-acids", "feed-additives"],
        ],
    )

    add_data_sheet(
        wb, "PACKAGING_TYPES",
        ["name", "slug"],
        ["REQUIRED — Packaging name", "REQUIRED — Unique slug"],
        [
            ["25 kg bag", "25-kg-bag"],
            ["Big bag 1000 kg", "big-bag-1000-kg"],
            ["Bulk", "bulk"],
            ["50 lb bag (lysine)", "50-lb-bag-lysine"],
            ["Tote 2000 lb", "tote-2000-lb"],
        ],
    )

    add_data_sheet(
        wb, "PARAMETERS",
        ["label", "slug", "type"],
        ["REQUIRED — Parameter name", "REQUIRED — Unique slug", "OPTIONAL — E.g. physical, chemical"],
        [
            ["Moisture", "moisture", "physical"],
            ["pH", "ph", "chemical"],
            ["Purity (lysine)", "purity-lysine", "chemical"],
        ],
    )

    add_data_sheet(
        wb, "TEST_METHODS",
        ["label", "slug"],
        ["REQUIRED — Test method name", "REQUIRED — Unique slug"],
        [["AOAC 930.15", "aoac-930-15"], ["ISO 6491", "iso-6491"], ["USP 891", "usp-891"]],
    )

    add_data_sheet(
        wb, "MEASURE_UNITS",
        ["label", "slug", "notation"],
        ["REQUIRED — Unit name", "REQUIRED — Unique slug", "OPTIONAL — Symbol. E.g. %"],
        [["Percent", "percent", "%"], ["Parts per million", "ppm", "ppm"]],
    )

    add_data_sheet(
        wb, "NUTRITIONAL_PARAMETERS",
        ["label", "slug", "notation"],
        ["REQUIRED — Component name", "REQUIRED — Unique slug", "OPTIONAL — Abbreviation. E.g. P"],
        [
            ["Phosphorus (P)", "phosphorus-p", "P"],
            ["Calcium (Ca)", "calcium-ca", "Ca"],
            ["Lysine", "lysine", "Lys"],
        ],
    )

    add_data_sheet(
        wb, "HANDLING_SPECS",
        ["label", "slug", "requirement"],
        ["REQUIRED — Spec name", "REQUIRED — Unique slug", "REQUIRED — Short detail (max 100 chars)"],
        [
            ["Store in dry place", "store-dry", "Keep away from moisture"],
            ["Temperature", "temperature", "Store below 35°C"],
            ["Keep sealed", "keep-sealed", "Reseal after opening"],
        ],
    )

    add_data_sheet(
        wb, "TYPICAL_APPLICATIONS",
        ["label", "slug", "description"],
        ["REQUIRED — Application name", "REQUIRED — Unique slug", "OPTIONAL — Extended description"],
        [
            ["Poultry feed", "poultry-feed", "Complete feeds and premixes for broilers and layers."],
            ["Swine feed", "swine-feed", ""],
            ["Aqua feed", "aqua-feed", "Fish and shrimp feed formulations."],
        ],
    )

    # --- PRODUCTS (one row per product; lists use pipe | as separator) ---

    product1_volume_tiers = (
        "1|Small orders|1|49|percentage|0||"
        "1|Volume|49||percentage|12||"
        "2|Tote single|1|3|fixed_price||1180|"
        "2|Tote bulk|3||fixed_price||1050"
    )

    add_data_sheet(
        wb, "PRODUCTS",
        [
            "slug", "name", "grade", "base_price_ref", "description",
            "status", "stock_status", "availability", "lead_time_days", "max_lead_time_days",
            "origin_address", "shelf_life_template", "market_trends_link",
            "tds_document_url", "sds_document_url", "coa_document_url",
            "category_slugs", "handling_spec_slugs", "application_slugs", "related_product_slugs",
            "packaging", "volume_tiers", "nutritional_analysis", "specifications",
        ],
        [
            "REQUIRED — Unique product identifier",
            "REQUIRED — Product name",
            "OPTIONAL — Grade / concentration",
            "OPTIONAL — Reference price (decimal)",
            "OPTIONAL — Description (HTML allowed)",
            "REQUIRED — draft, published, or archived",
            "OPTIONAL — in_stock, out_of_stock, backorder, or call",
            "OPTIONAL — Customer-facing availability text",
            "OPTIONAL — Lead time in days (integer)",
            "OPTIONAL — Max lead time in days",
            "OPTIONAL — Origin address",
            "OPTIONAL — Shelf life",
            "OPTIONAL — Market trends URL",
            "OPTIONAL — TDS PDF URL or path",
            "OPTIONAL — SDS PDF URL or path",
            "OPTIONAL — COA PDF URL or path",
            "OPTIONAL — Pipe-separated category slugs from CATEGORIES. E.g. phosphates|feed-additives",
            "OPTIONAL — Pipe-separated slugs from HANDLING_SPECS. E.g. store-dry|temperature",
            "OPTIONAL — Pipe-separated slugs from TYPICAL_APPLICATIONS. E.g. poultry-feed|swine-feed",
            "OPTIONAL — Pipe-separated product slugs (must exist in DB or earlier rows). E.g. urea-feed-grade",
            "OPTIONAL — 4 fields per item. presentation_index|packaging_type_slug|quantity_per_pallet|base_price_per_unit",
            "OPTIONAL — 7 fields per item. presentation_index|tier_name|min|max|pricing_mode|discount_pct|fixed_price. Empty max on last tier = unlimited",
            "OPTIONAL — 3 fields per item. nutritional_parameter_slug|value|measure_unit_slug",
            "OPTIONAL — 4 fields per item. parameter_slug|test_method_slug|specification|measure_unit_slug",
        ],
        [
            [
                "premium-lysine-hcl",
                "Premium Lysine HCl",
                "Feed grade 98.5%",
                620.00,
                "<p>High-purity lysine for aqua and livestock feeds. Example: new masters + 2 packagings with volume tiers.</p>",
                "published",
                "in_stock",
                "Immediate",
                7,
                14,
                "USA",
                "24 months",
                "",
                "",
                "",
                "",
                "amino-acids",
                "keep-sealed",
                "aqua-feed",
                "",
                "1|50-lb-bag-lysine|48|620.00|2|tote-2000-lb|1|850.00",
                product1_volume_tiers,
                "lysine|98.5|percent",
                "purity-lysine|usp-891|Min 98.5|percent",
            ],
            [
                "lysine-premix-blend",
                "Lysine Premix Blend",
                "18%",
                85.00,
                "<p>Premix using existing catalog slugs. One packaging, base price only, no volume tiers.</p>",
                "published",
                "in_stock",
                "Immediate",
                10,
                14,
                "USA",
                "12 months",
                "",
                "",
                "",
                "",
                "phosphates",
                "store-dry",
                "poultry-feed",
                "premium-lysine-hcl|urea-feed-grade",
                "1|25-kg-bag|40|85.00",
                "",
                "phosphorus-p|18.5|percent",
                "moisture|aoac-930-15|Max 12|percent",
            ],
        ],
    )

    wb.save(OUTPUT)
    print(f"Template generated: {OUTPUT}")


if __name__ == "__main__":
    main()
