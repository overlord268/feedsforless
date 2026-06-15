#!/usr/bin/env python3
"""
Generate the FeedsForLess product import Excel template (English, product-centric).
Run: python scripts/generate-product-import-template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "product-import-template-feedsforless.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1B4332")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DESC_FILL = PatternFill("solid", fgColor="D8F3DC")
DESC_FONT = Font(italic=True, color="495057", size=9)
TITLE_FONT = Font(bold=True, size=14, color="1B4332")
SUBTITLE_FONT = Font(bold=True, size=11, color="2D6A4F")
THIN = Side(style="thin", color="ADB5BD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_desc_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = DESC_FILL
        cell.font = DESC_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER


def auto_width(ws, min_width=12, max_width=60):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        length = min_width
        for cell in column_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = length


def add_data_sheet(wb, name, headers, descriptions, examples=None):
    ws = wb.create_sheet(name)
    for col, (header, desc) in enumerate(zip(headers, descriptions), 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=2, column=col, value=desc)
    style_header_row(ws, 1, len(headers))
    style_desc_row(ws, 2, len(headers))
    if examples:
        for row_idx, example in enumerate(examples, 3):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A3"
    auto_width(ws)
    return ws


def build_instructions_sheet(wb):
    ws = wb.active
    ws.title = "INSTRUCTIONS"

    lines = [
        ("FeedsForLess — Product Import Template", TITLE_FONT),
        ("", None),
        ("Purpose", SUBTITLE_FONT),
        ("Collect product catalog data from an external website in a structured format for import into FeedsForLess.", None),
        ("", None),
        ("Product-centric design (no separate master sheets)", SUBTITLE_FONT),
        ("Master catalog values (categories, packaging types, parameters, etc.) are defined inline while filling product rows.", None),
        ("During import, the system will create any referenced master record that does not exist yet, using the columns provided.", None),
        ("This matches how the admin product form works: you pick or create masters from within the product.", None),
        ("", None),
        ("Recommended fill order", SUBTITLE_FONT),
        ("1. Fill PRODUCTS — one row per product (general fields).", None),
        ("2. Fill PRODUCT_CATEGORIES — at least one row per product.", None),
        ("3. Fill remaining sheets as needed (packaging, specs, nutritional analysis, etc.).", None),
        ("4. Use the same sku value across all sheets to link rows to a product.", None),
        ("5. Review FIELD_REFERENCE for column details and allowed values.", None),
        ("", None),
        ("General rules", SUBTITLE_FONT),
        ("• Do not rename column headers (row 1) or remove the description row (row 2).", None),
        ("• sku must be unique per product and is the key linking all sheets.", None),
        ("• Required fields are marked REQUIRED in row 2.", None),
        ("• Master auto-create: if a category, packaging type, parameter, etc. is referenced by label/name and does not exist, it will be created on import.", None),
        ("• Optional master columns (type, notation, requirement, description) only need to be filled once; later rows can omit them if the label/name matches.", None),
        ("• Product description supports basic HTML (<b>, <i>, <ul>, <li>, <p>).", None),
        ("• TDS / SDS / COA: provide a public PDF URL from the source site, or a relative path if the file is already available.", None),
        ("• PDF files can also be delivered separately as: {SKU}_TDS.pdf, {SKU}_SDS.pdf, {SKU}_COA.pdf", None),
        ("", None),
        ("Sheets in this file", SUBTITLE_FONT),
        ("  • PRODUCTS — Main product data (1 row = 1 product).", None),
        ("  • PRODUCT_CATEGORIES — Category assignments (min. 1 per product).", None),
        ("  • PRODUCT_PACKAGING — Presentations and base pricing (multiple rows per product).", None),
        ("  • PRODUCT_VOLUME_TIERS — Volume discount tiers per packaging row.", None),
        ("  • PRODUCT_NUTRITIONAL_ANALYSIS — Nutritional analysis rows.", None),
        ("  • PRODUCT_SPECIFICATIONS — Technical specification rows.", None),
        ("  • PRODUCT_HANDLING_SPECS — Handling / storage requirements.", None),
        ("  • PRODUCT_APPLICATIONS — Typical applications.", None),
        ("  • PRODUCT_RELATED — Related products (cross-sell).", None),
        ("", None),
        ("Allowed enum values", SUBTITLE_FONT),
        ("status: draft | published | archived", None),
        ("stock_status: in_stock | out_of_stock | backorder | call", None),
    ]

    for row_idx, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 105


def build_reference_sheet(wb):
    ws = wb.create_sheet("FIELD_REFERENCE")
    headers = ["Sheet", "Column", "Required", "Type", "Description / allowed values"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    rows = [
        ["PRODUCTS", "sku", "Yes", "Text", "Unique product code. Key for all related sheets."],
        ["PRODUCTS", "name", "Yes", "Text", "Commercial product name."],
        ["PRODUCTS", "slug", "No", "Text", "URL slug. Auto-generated from name if empty."],
        ["PRODUCTS", "grade", "No", "Text", "Grade or concentration. E.g. 18.50%"],
        ["PRODUCTS", "base_price_ref", "No", "Number", "Reference price (decimal)."],
        ["PRODUCTS", "description", "No", "HTML", "Detailed product description."],
        ["PRODUCTS", "status", "Yes", "Enum", "draft | published | archived"],
        ["PRODUCTS", "stock_status", "No", "Enum", "in_stock | out_of_stock | backorder | call"],
        ["PRODUCTS", "availability", "No", "Text", "Customer-facing availability text. E.g. Immediate"],
        ["PRODUCTS", "lead_time_days", "No", "Integer", "Estimated lead time in days."],
        ["PRODUCTS", "max_lead_time_days", "No", "Integer", "Maximum lead time in days."],
        ["PRODUCTS", "origin_address", "No", "Text", "Origin / source address."],
        ["PRODUCTS", "shelf_life_template", "No", "Text", "Shelf life. E.g. 24 months"],
        ["PRODUCTS", "market_trends_link", "No", "URL", "External market trends link."],
        ["PRODUCTS", "tds_document_url", "No", "URL/path", "Technical Data Sheet (PDF)."],
        ["PRODUCTS", "sds_document_url", "No", "URL/path", "Safety Data Sheet (PDF)."],
        ["PRODUCTS", "coa_document_url", "No", "URL/path", "Certificate of Analysis (PDF)."],
        ["PRODUCT_CATEGORIES", "sku", "Yes", "Text", "Product SKU."],
        ["PRODUCT_CATEGORIES", "category_label", "Yes", "Text", "Category display name. Created if missing."],
        ["PRODUCT_CATEGORIES", "category_slug", "No", "Text", "URL slug. Auto-generated from label if empty."],
        ["PRODUCT_CATEGORIES", "parent_category_slug", "No", "Text", "Parent category slug for hierarchy. Empty = root."],
        ["PRODUCT_PACKAGING", "sku", "Yes", "Text", "Product SKU."],
        ["PRODUCT_PACKAGING", "presentation_index", "Yes", "Integer", "Presentation number (1, 2, 3…) used to link volume tiers."],
        ["PRODUCT_PACKAGING", "packaging_type_name", "Yes", "Text", "Packaging type name. Created if missing."],
        ["PRODUCT_PACKAGING", "quantity_per_pallet", "Yes", "Integer", "Units per pallet (min. 1)."],
        ["PRODUCT_PACKAGING", "base_price_per_unit", "Yes", "Number", "Base price per unit for this presentation."],
        ["PRODUCT_VOLUME_TIERS", "sku", "Yes", "Text", "Product SKU."],
        ["PRODUCT_VOLUME_TIERS", "presentation_index", "Yes", "Integer", "Must match PRODUCT_PACKAGING.presentation_index."],
        ["PRODUCT_VOLUME_TIERS", "tier_name", "Yes", "Text", "Tier label. E.g. Tier 1"],
        ["PRODUCT_VOLUME_TIERS", "min_quantity", "Yes", "Integer", "Minimum quantity for this tier."],
        ["PRODUCT_VOLUME_TIERS", "max_quantity", "No", "Integer", "Maximum quantity. Empty = unlimited."],
        ["PRODUCT_VOLUME_TIERS", "discount_percentage", "Yes", "Number", "Discount 0–100."],
        ["PRODUCT_NUTRITIONAL_ANALYSIS", "sku", "Yes", "Text", "Product SKU."],
        ["PRODUCT_NUTRITIONAL_ANALYSIS", "nutritional_parameter_label", "Yes", "Text", "Nutritional parameter. Created if missing."],
        ["PRODUCT_NUTRITIONAL_ANALYSIS", "nutritional_parameter_notation", "No", "Text", "Abbreviation when creating parameter. E.g. P"],
        ["PRODUCT_NUTRITIONAL_ANALYSIS", "value", "No", "Text", "Reported value. E.g. 18.5"],
        ["PRODUCT_NUTRITIONAL_ANALYSIS", "measure_unit_notation", "No", "Text", "Unit symbol. Created if missing (label defaults to notation)."],
        ["PRODUCT_SPECIFICATIONS", "sku", "Yes", "Text", "Product SKU."],
        ["PRODUCT_SPECIFICATIONS", "parameter_label", "Yes", "Text", "Technical parameter. Created if missing."],
        ["PRODUCT_SPECIFICATIONS", "parameter_type", "No", "Text", "Optional type when creating parameter. E.g. physical"],
        ["PRODUCT_SPECIFICATIONS", "test_method_label", "Yes", "Text", "Test method. Created if missing."],
        ["PRODUCT_SPECIFICATIONS", "specification", "Yes", "Text", "Required value or range. E.g. Max 12"],
        ["PRODUCT_SPECIFICATIONS", "measure_unit_notation", "Yes", "Text", "Unit symbol. Created if missing."],
        ["PRODUCT_HANDLING_SPECS", "sku", "Yes", "Text", "Product SKU."],
        ["PRODUCT_HANDLING_SPECS", "handling_spec_label", "Yes", "Text", "Handling spec name. Created if missing."],
        ["PRODUCT_HANDLING_SPECS", "handling_spec_requirement", "No", "Text", "Detail when creating spec (max 100 chars). Defaults to label if empty."],
        ["PRODUCT_APPLICATIONS", "sku", "Yes", "Text", "Product SKU."],
        ["PRODUCT_APPLICATIONS", "application_label", "Yes", "Text", "Typical application. Created if missing."],
        ["PRODUCT_APPLICATIONS", "application_description", "No", "Text", "Optional description when creating application."],
        ["PRODUCT_RELATED", "sku", "Yes", "Text", "Main product SKU."],
        ["PRODUCT_RELATED", "related_sku", "Yes", "Text", "Related product SKU (must exist in PRODUCTS)."],
    ]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    auto_width(ws)


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_instructions_sheet(wb)
    build_reference_sheet(wb)

    add_data_sheet(
        wb, "PRODUCTS",
        [
            "sku", "name", "slug", "grade", "base_price_ref", "description",
            "status", "stock_status", "availability", "lead_time_days", "max_lead_time_days",
            "origin_address", "shelf_life_template", "market_trends_link",
            "tds_document_url", "sds_document_url", "coa_document_url",
        ],
        [
            "REQUIRED — Unique product code",
            "REQUIRED — Product name",
            "OPTIONAL — URL slug; auto-generated if empty",
            "OPTIONAL — Grade / concentration",
            "OPTIONAL — Reference price (decimal)",
            "OPTIONAL — Description (HTML allowed)",
            "REQUIRED — draft | published | archived",
            "OPTIONAL — in_stock | out_of_stock | backorder | call",
            "OPTIONAL — Customer-facing availability text",
            "OPTIONAL — Lead time in days (integer)",
            "OPTIONAL — Max lead time in days",
            "OPTIONAL — Origin address",
            "OPTIONAL — Shelf life",
            "OPTIONAL — Market trends URL",
            "OPTIONAL — TDS PDF URL or path",
            "OPTIONAL — SDS PDF URL or path",
            "OPTIONAL — COA PDF URL or path",
        ],
        [[
            "DCP-185", "Dicalcium Feed Phosphate", "dicalcium-feed-phosphate", "18.50%", 450.00,
            "<p>Feed-grade calcium and phosphorus supplement for livestock.</p>",
            "published", "in_stock", "Immediate", 10, 14,
            "USA", "24 months", "",
            "https://example.com/docs/DCP-TDS.pdf",
            "https://example.com/docs/DCP-SDS.pdf", "",
        ]],
    )

    add_data_sheet(
        wb, "PRODUCT_CATEGORIES",
        ["sku", "category_label", "category_slug", "parent_category_slug"],
        [
            "REQUIRED — Product SKU",
            "REQUIRED — Category name (auto-created if missing)",
            "OPTIONAL — URL slug; auto-generated from label if empty",
            "OPTIONAL — Parent category slug for hierarchy",
        ],
        [["DCP-185", "Phosphates", "phosphates", ""]],
    )

    add_data_sheet(
        wb, "PRODUCT_PACKAGING",
        ["sku", "presentation_index", "packaging_type_name", "quantity_per_pallet", "base_price_per_unit"],
        [
            "REQUIRED — Product SKU",
            "REQUIRED — Index 1, 2, 3… (links to volume tiers)",
            "REQUIRED — Packaging type name (auto-created if missing)",
            "REQUIRED — Units per pallet (min. 1)",
            "REQUIRED — Base price per unit",
        ],
        [["DCP-185", 1, "25 kg bag", 40, 12.50], ["DCP-185", 2, "Big bag 1000 kg", 20, 480.00]],
    )

    add_data_sheet(
        wb, "PRODUCT_VOLUME_TIERS",
        ["sku", "presentation_index", "tier_name", "min_quantity", "max_quantity", "discount_percentage"],
        [
            "REQUIRED — Product SKU",
            "REQUIRED — Must match PRODUCT_PACKAGING.presentation_index",
            "REQUIRED — Tier name",
            "REQUIRED — Minimum quantity",
            "OPTIONAL — Maximum quantity (empty = unlimited)",
            "REQUIRED — Discount % (0–100)",
        ],
        [["DCP-185", 1, "Tier 1", 1, 99, 0], ["DCP-185", 1, "Tier 2", 100, None, 5]],
    )

    add_data_sheet(
        wb, "PRODUCT_NUTRITIONAL_ANALYSIS",
        ["sku", "nutritional_parameter_label", "nutritional_parameter_notation", "value", "measure_unit_notation"],
        [
            "REQUIRED — Product SKU",
            "REQUIRED — Parameter name (auto-created if missing)",
            "OPTIONAL — Abbreviation when creating parameter",
            "OPTIONAL — Reported value",
            "OPTIONAL — Unit symbol (auto-created if missing)",
        ],
        [["DCP-185", "Phosphorus (P)", "P", "18.5", "%"], ["DCP-185", "Calcium (Ca)", "Ca", "21", "%"]],
    )

    add_data_sheet(
        wb, "PRODUCT_SPECIFICATIONS",
        ["sku", "parameter_label", "parameter_type", "test_method_label", "specification", "measure_unit_notation"],
        [
            "REQUIRED — Product SKU",
            "REQUIRED — Parameter name (auto-created if missing)",
            "OPTIONAL — Type when creating parameter. E.g. physical",
            "REQUIRED — Test method (auto-created if missing)",
            "REQUIRED — Value or range",
            "REQUIRED — Unit symbol (auto-created if missing)",
        ],
        [["DCP-185", "Moisture", "physical", "AOAC 930.15", "Max 12", "%"]],
    )

    add_data_sheet(
        wb, "PRODUCT_HANDLING_SPECS",
        ["sku", "handling_spec_label", "handling_spec_requirement"],
        [
            "REQUIRED — Product SKU",
            "REQUIRED — Handling spec name (auto-created if missing)",
            "OPTIONAL — Requirement detail (max 100 chars); defaults to label",
        ],
        [["DCP-185", "Store in dry place", "Keep away from moisture"]],
    )

    add_data_sheet(
        wb, "PRODUCT_APPLICATIONS",
        ["sku", "application_label", "application_description"],
        [
            "REQUIRED — Product SKU",
            "REQUIRED — Application name (auto-created if missing)",
            "OPTIONAL — Description when creating application",
        ],
        [["DCP-185", "Poultry feed", "Complete feeds and premixes for broilers and layers."], ["DCP-185", "Swine feed", ""]],
    )

    add_data_sheet(
        wb, "PRODUCT_RELATED",
        ["sku", "related_sku"],
        ["REQUIRED — Main product SKU", "REQUIRED — Related product SKU"],
        [["DCP-185", "MCP-227"]],
    )

    wb.save(OUTPUT)
    print(f"Template generated: {OUTPUT}")


if __name__ == "__main__":
    main()
